#Author Mporras
#Date 13/sep/2018

import ast
import arrow
import os
from os import listdir
from os.path import isfile, join
from openpyxl import Workbook

def readFolder(executionMethod, reportType, isDev, path):
    path += "car/"
    onlyfiles = [f for f in listdir(path) if isfile(join(path, f))]
    filePlates = []
    cars = []    
    titles = ["Video", "Placa", "Porcentaje", "Tipo de placa", "Correcta/Incorrecta", "Path", "Valor correcto placa", "Valor placa no encontrada"]
    separators = ["-------------", "-------------", "-------------", "-------------", "-------------", "-------------"]    
    for fileName in onlyfiles:                
        readFileResult = readFile(executionMethod, reportType, isDev, path + fileName)        
        if readFileResult != None:
            cars.append(readFileResult[0])
            filePlates.append(readFileResult[1]) 
    
    if reportType == "--Early_Morning":        
        generateReport("./reports/Early_Morning/" + arrow.now().format('YYYY-MM-DD') + "-Early_Morning.xlsx", titles, separators, cars, filePlates)
    elif reportType == "--Morning":        
        generateReport("./reports/Morning/" + arrow.now().format('YYYY-MM-DD') + "-Morning.xlsx", titles, separators, cars, filePlates)
    elif reportType == "--Afternoon" :        
        generateReport("./reports/Afternoon/" + arrow.now().format('YYYY-MM-DD') + "-Afternoon.xlsx", titles, separators, cars, filePlates)
    elif reportType == "--Night":        
        generateReport("./reports/Night/" + arrow.now().format('YYYY-MM-DD') + "-Night.xlsx", titles, separators, cars, filePlates)
    else:
        ### pending for development
        generateGeneralReport(cars, filePlates)

pass

def readFolder_Car_NoCar(executionMethod, reportType, isDev, path):    
    cars = []  
    notCars = []

    alternativePath = path + "car/"
    if os.path.exists(alternativePath):
        onlyfiles = [f for f in listdir(alternativePath) if isfile(join(alternativePath, f))]                          
        for fileName in onlyfiles:                
            readFileResult = readFile(executionMethod, reportType, isDev, alternativePath + fileName)        
            if readFileResult != None:                
                cars.append(readFileResult[0])                        

    alternativePath = path + "notCar/"
    if os.path.exists(alternativePath):
        onlyfiles = [f for f in listdir(alternativePath) if isfile(join(alternativePath, f))]        
        for fileName in onlyfiles:                
            readFileResult = readFile(executionMethod, reportType, isDev, alternativePath + fileName)        
            if readFileResult != None:                
                notCars.append(readFileResult[0]) 
    
    titles = ["Car/notCar", "Path", "Video", "Car", "notCar"]
    separators = ["-------------", "-------------", "-------------"]

    if reportType == "--Early_Morning":                
        generateReport_Car_NoCar("./reports/Early_Morning/" + arrow.now().format('YYYY-MM-DD') + "-Early_Morning_car_notCar.xlsx", titles, separators, cars, notCars)            
    elif reportType == "--Morning":                
        generateReport_Car_NoCar("./reports/Morning/" + arrow.now().format('YYYY-MM-DD') + "-Morning_car_notCar.xlsx", titles, separators, cars, notCars)                    
    elif reportType == "--Afternoon" :        
        generateReport_Car_NoCar("./reports/Afternoon/" + arrow.now().format('YYYY-MM-DD') + "-Afternoon_car_notCar.xlsx", titles, separators, cars, notCars)                            
    elif reportType == "--Night":        
        generateReport_Car_NoCar("./reports/Night/" + arrow.now().format('YYYY-MM-DD') + "-Night_car_notCar.xlsx", titles, separators, cars, notCars)                                    

    pass

def readFile(executionMethod, reportType, isDev, path):
    i = 0
    devFlag = True if isDev == "--dev" else False
    cars = []
    plates = []

    if path.find(".DS_Store") < 0:        
        file = open(path, 'r')
        for line in file:
            if devFlag and i < 3:
                cars.append(line)
            else:
                plates.append(line)
            i += 1  
        return (cars, plates)
          
    pass

def generateReport(path, titles, separators, cars, plates):
    wb = Workbook()
    ws = wb.active
    
    ws.append(titles)
    i = 0

    for value in plates:            
        if len(value) == 0:       
            print("File not found")                       
            ws.append([cars[i][2] + ".mp4", '', '', '', '', cars[i][1].replace("\n", ".mp4"), 0, 1])

        for plate in value:            
            plate = plate.split(",")               
            plate[len(plate) - 1] = plate[len(plate) - 1].split(".mp4")[0] + ".mp4"            
            isCorrect = plate[4].strip() if len(plate) > 4 else "Incorrect"
            if isCorrect == "Correct":
                plate.append(1)
                plate.append(0)
            elif isCorrect == "Incorrect":
                plate.append(0)
                plate.append(0)            

            ws.append(plate)

        ws.append(separators)
        i += 1

    wb.save(path)  
pass

def generateReport_Car_NoCar(path, titles, separatos, cars, notCars):
    wb = Workbook()
    ws = wb.active

    ws.append(titles)

    for car in cars:
        car.append(1)
        car.append(0)
        ws.append(car)
        ws.append(separatos)

    for notCar in notCars:
        notCar.append(0)
        notCar.append(1)
        ws.append(notCar)
        ws.append(separatos)

    wb.save(path)
    pass

# def readMultiplePaths(paths, isDev):
#     # [path += "/Early_Morning/car/", path += "/Morning/car/", path += "/Afternoon/car/", path += "/Night/car/"]
#     early_Morning_Data = getFolderData(isDev, paths[0])    
#     Morning_Data = getFolderData(isDev, paths[1])
#     Afternoon_Data = getFolderData(isDev, paths[2])
#     Night_Data = getFolderData(isDev, paths[3])    

#     generateGeneralReport(early_Morning_Data[1], Morning_Data[1], Afternoon_Data[1], Night_Data[1])
# pass

# def getFolderData(isDev, path):
#     onlyfiles = [f for f in listdir(path) if isfile(join(path, f))]
#     filePlates = []
#     cars = []    
    
#     for fileName in onlyfiles:                
#         readFileResult = getFileData(isDev, path + fileName)        
#         if readFileResult != None:
#             cars.append(readFileResult[0])
#             filePlates.append(readFileResult[1]) 

#     return (cars, filePlates)
# pass

# def getFileData(isDev, path):
#     i = 0
#     devFlag = True if isDev == "--dev" else False
#     cars = []
#     plates = []

#     if path.find(".DS_Store") < 0:        
#         file = open(path, 'r')
#         for line in file:
#             if devFlag and i < 3:
#                 cars.append(line)
#             else:
#                 plates.append(line)
#             i += 1  
#         return (cars, plates)
          
#     pass

# def generateGeneralReport(early_Morning_Data, Morning_Data, Afternoon_Data, Night_Data):
#     titles = ["Video", "Placa", "Porcentaje", "Tipo de placa", "Correcta/Incorrecta", "Path", "Valor correcto placa", "Valor placa no encontrada" ,"origen"]
#     separators = ["-------------", "-------------", "-------------", "-------------", "-------------", "-------------", "", "", "-------------"]

#     wb = Workbook()
#     ws = wb.active
    
#     ws.append(titles)
    
#     i = 0
#     plate = []

#     for early_value in early_Morning_Data:                    
#         if len(early_value) == 3:
#             plate = [early_value[2], "", "0%", "not plate", "Incorrect", early_value[1], 0, 1, "Early Morning"]
#         else:                        
#             i = 0
#             for early_morning in morning_value:
#                 if i > 3:
#                     plate = early_morning.split(",")      
#                     plate.append(1)
#                     plate.append(0)
#                     plate.append("Early_Morning")
#                 i += 1    
#         ws.append(plate)

    
#     for morning_value in Morning_Data:                    
#         if len(morning_value) == 3:
#             plate = [morning_value[2], "", "0%", "not plate", "Incorrect", morning_value[1], 0, 1, "Morning"]
#         else:
#             i = 0
#             for morning in morning_value:
#                 if i > 3:
#                     plate = morning.split(",")
#                     plate.append(1)
#                     plate.append(0)
#                     plate.append("Morning")                
#                 i += 1                        
#         ws.append(plate)


#     for afternoon_value in Afternoon_Data:                    
#         if len(afternoon_value) == 3:
#             plate = [afternoon_value[2], "", "0%", "not plate", "Incorrect", afternoon_value[1], 0, 1, "Afternoon"]
#         else:
#             i = 0
#             for afternoon in afternoon_value:
#                 if i > 3:
#                     plate = morning.split(",")
#                     plate.append(1)
#                     plate.append(0)
#                     plate.append("Afternoon")                
#                 i += 1    
#         ws.append(plate)
            

#     for night_value in Night_Data:                    
#         if len(night_value) == 3:
#             plate = [night_value[2], "", "0%", "not plate", "Incorrect", night_value[1], 0, 1, "Night"]
#         else:
#             i = 0
#             for night in night_value:
#                 if i > 3:
#                     plate = morning.split(",")
#                     plate.append(1)
#                     plate.append(0)
#                     plate.append("Night")                
#                 i += 1
#         ws.append(plate)


#     wb.save("./reports/General/" + arrow.now().format('YYYY-MM-DD') + "-General.xlsx") 

# pass